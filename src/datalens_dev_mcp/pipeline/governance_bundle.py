from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from datalens_dev_mcp.pipeline.route_registry import decide_registered_route

sys.dont_write_bytecode = True


SCHEMA_VERSION = "2026-05-17.requirements_s2t_intake.v1"
MAX_SAMPLE_PROFILE_ROWS = 10000
NULL_RATIO_QUALITY_FAIL = 0.2
SMALL_SAMPLE_FAIL = 30
SMALL_SAMPLE_WARN = 100
FRESHNESS_DEFAULT_HOURS = {
    "hourly": 6,
    "daily": 48,
    "weekly": 336,
    "unknown": 168,
}
SUPPORT_STATUSES = {
    "supported",
    "supported_with_assumption",
    "methodology_pending",
    "data_missing",
    "source_not_integrated",
    "quality_risk",
}
STATUS_TO_AVAILABILITY = {
    "supported": "available",
    "supported_with_assumption": "available",
    "methodology_pending": "unavailable_methodology_pending",
    "data_missing": "unavailable_missing_data",
    "source_not_integrated": "unavailable_missing_data",
    "quality_risk": "unavailable_invalid_data",
}
STATUS_TO_DECISION_STATUS = {
    "supported": "ready",
    "supported_with_assumption": "draft_with_assumptions",
    "methodology_pending": "needs_clarification",
    "data_missing": "blocked_missing_context",
    "source_not_integrated": "blocked_missing_context",
    "quality_risk": "blocked_data_quality",
}
STATE_TITLES = {
    "available": "NO DATA",
    "unavailable_missing_data": "SOURCE MISSING",
    "unavailable_invalid_data": "QUALITY ISSUE",
    "unavailable_methodology_pending": "METHODOLOGY PENDING",
    "filtered_empty": "FILTERED OUT",
    "true_zero": "NO DATA",
}
MAP_FAMILIES = {"native_map_geo_widget", "native_map_widget", "symbol_map"}
MARKDOWN_FAMILIES = {
    "md_section_header",
    "md_dashboard_owner",
    "md_contact_block",
    "md_requirements_link_block",
    "md_methodology_block",
    "md_source_notes",
    "md_methodology_cut",
    "md_methodology_accordion",
    "md_doc_tabs",
    "markdown_explainer_block",
}
CONTROL_FAMILIES = {
    "selector_family_static",
    "selector_family_dynamic",
    "global_top_bar_selector_row",
    "local_section_selector_row",
    "metric_toggle_selector",
    "date_range_selector",
    "week_selector",
    "single_select_dropdown",
    "multi_select_dropdown",
    "search_selector",
    "reset_clear_pattern",
}
ALLOWED_VISUAL_FAMILIES = {
    "kpi_value_only",
    "kpi_value_delta",
    "kpi_value_sparkline",
    "kpi_value_delta_sparkline",
    "line_chart",
    "vertical_bar_time_bucket",
    "combo_time_series_combo",
    "funnel_snapshot",
    "bump_chart",
    "horizontal_bar",
    "grouped_bar",
    "stacked_100",
    "heatmap",
    "waterfall",
    "histogram",
    "box_plot",
    "scatter",
    "bubble",
    "pie",
    "donut",
    "treemap",
    "native_map_geo_widget",
    "native_map_widget",
    "symbol_map",
    "table_node",
    *MARKDOWN_FAMILIES,
    *CONTROL_FAMILIES,
}


def expected_route_for_family(family: str) -> str:
    return decide_registered_route(family).route


def expected_tabs_for_route(route: str) -> list[str]:
    if route == "editor_table":
        return ["meta.json", "params.js", "sources.js", "prepare.js", "config.js"]
    if route == "editor_markdown":
        return ["meta.json", "params.js", "sources.js", "prepare.js"]
    if route == "editor_js_control":
        return ["meta.json", "params.js", "sources.js", "controls.js"]
    if route == "editor_advanced":
        return ["meta.json", "params.js", "sources.js", "prepare.js"]
    if route == "wizard_native":
        return []
    return []


def slugify(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "_", value.strip().lower())
    return re.sub(r"_+", "_", value).strip("_") or "item"


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_datetime_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None
    else:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def parse_as_of(value: str | None) -> datetime | None:
    if not value:
        return None
    parsed = parse_datetime_value(value)
    if not parsed:
        raise RuntimeError(f"--as-of must be ISO datetime, got {value!r}")
    return parsed


def compact_datetime(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_pdf_text(path: Path) -> str:
    executable = shutil.which("pdftotext")
    if not executable:
        raise RuntimeError("pdftotext is required for direct PDF intake and was not found on PATH.")
    result = subprocess.run([executable, str(path), "-"], check=False, capture_output=True, text=True)
    if result.returncode != 0:
        detail = result.stderr.strip() or result.stdout.strip() or "pdftotext failed without details"
        raise RuntimeError(f"pdftotext failed for {path}: {detail}")
    text = result.stdout.strip()
    if not text:
        raise RuntimeError(f"pdftotext produced no text for {path}; OCR is out of scope for this pipeline.")
    return text


def first_match(text: str, patterns: list[str], default: str = "unknown") -> str:
    for pattern in patterns:
        match = re.search(pattern, text, re.I | re.M)
        if match:
            return " ".join(match.group(1).strip().split())
    return default


def parse_requirement_text(text: str, source_name: str) -> dict[str, Any]:
    compact = text[:8000]
    dashboard_name = first_match(
        compact,
        [
            r"Название дашборда\s*\n+([^\n]+)",
            r"Dashboard name\s*[:\n]+([^\n]+)",
            r"^([A-Z][A-Za-z0-9 /_-]+ Dashboard)\b",
        ],
        Path(source_name).stem,
    )
    update_frequency = first_match(
        compact,
        [r"Частота обновления\s*\n+([^\n]+)", r"Update frequency\s*[:\n]+([^\n]+)", r"\b(daily|hourly|weekly|monthly)\b"],
        "missing",
    )
    objective = first_match(
        compact,
        [r"Цель разработки\s*\n+(.+?)(?:\n\s*\n|Dashboard|Business|Data Source)", r"Objective\s*[:\n]+(.+?)(?:\n\s*\n|Business|Data Source)"],
        "Extracted requirement PDF needs manual objective review.",
    )
    metrics = infer_requirement_metrics(compact)
    if not metrics:
        metrics = [
            {
                "metric_id": "primary_metric",
                "name": "Primary requested metric",
                "business_question": "What is the main dashboard signal?",
                "analytical_task": "kpi_status",
                "required_fields": [],
                "support_status": "methodology_pending",
                "expected_family": "kpi_value_sparkline",
                "metric_semantics": default_metric_semantics(),
                "missing_context": ["metric definition", "source field mapping"],
            }
        ]
    return {
        "dashboard_name": dashboard_name,
        "contact": first_match(compact, [r"Main contact[^\n]*\n+([^\n]+)", r"Контакт[^\n]*\n+([^\n]+)"], "missing"),
        "process": first_match(compact, [r"Process[^\n]*\n+([^\n]+)", r"Процесс[^\n]*\n+([^\n]+)"], "missing"),
        "lifetime": first_match(compact, [r"Lifetime[^\n]*\n+([^\n]+)", r"Срок жизни[^\n]*\n+([^\n]+)"], "missing"),
        "update_frequency": update_frequency,
        "objective": objective,
        "background": "Extracted from requirement PDF by deterministic heading heuristics.",
        "business_value": first_match(compact, [r"Business Sense\s+(.+?)(?:\n\s*\n|Data Sources|Metrics)", r"Бизнес[^\n]*\n+(.+?)(?:\n\s*\n|Источники|Метрики)"], "missing"),
        "audience": ["dashboard customer", "process owner"],
        "decision_action": "Use the dashboard to monitor the requested process and decide follow-up action.",
        "data_sources": [],
        "metrics": metrics,
        "visual_requirements": [],
        "sample_sketches": [],
        "open_questions": ["Confirm extracted fields and metric methods before deployment."],
        "out_of_scope": [],
    }


def infer_requirement_metrics(text: str) -> list[dict[str, Any]]:
    candidates = [
        ("lead_time", "Lead Time", "time_trend", "line_chart", ["created_dttm", "resolution_dttm"]),
        ("id_placeholder", "Predictability", "id_placeholder", "box_plot", []),
        ("backlog_wip", "Backlog and WIP", "matrix", "heatmap", ["status_nm"]),
        ("completion", "Completion", "flow_or_process", "funnel_snapshot", []),
        ("conversion", "Conversion", "flow_or_process", "funnel_snapshot", []),
        ("kwh", "kWh", "time_trend", "line_chart", []),
        ("api_uptime", "API uptime", "kpi_status", "kpi_value_sparkline", []),
        ("events", "Events over time", "time_trend", "line_chart", []),
        ("map", "Geo signal", "geo_pattern", "native_map_geo_widget", []),
    ]
    lower = text.lower()
    metrics: list[dict[str, Any]] = []
    for metric_id, name, task, family, fields in candidates:
        if metric_id.replace("_", " ") in lower or name.lower() in lower:
            metrics.append(
                {
                    "metric_id": metric_id,
                    "name": name,
                    "business_question": f"What is the state of {name.lower()}?",
                    "analytical_task": task,
                    "required_fields": fields,
                    "expected_family": family,
                    "metric_semantics": default_metric_semantics(),
                    "missing_context": [],
                }
            )
    return metrics


def parse_s2t_text(text: str, source_name: str) -> dict[str, Any]:
    compact = text[:12000]
    fields = []
    for line in compact.splitlines():
        stripped = " ".join(line.split())
        if re.match(r"^[A-Za-z][A-Za-z0-9_]{2,}\s+", stripped):
            name = stripped.split()[0]
            if "_" in name or name.endswith(("id", "rk", "cd", "nm", "dttm", "dt")):
                fields.append({"name": name, "type": "unknown", "flags": []})
    fields = dedupe_fields(fields)
    datetime_fields = [field["name"] for field in fields if field["name"].lower().endswith(("dt", "dttm", "date"))]
    return {
        "contract_id": slugify(Path(source_name).stem),
        "table_name": first_match(compact, [r"(?:iceberg|clickhouse):\s*([A-Za-z0-9_.]+)", r"Таблица\s*[:\n]+([A-Za-z0-9_.]+)"], slugify(Path(source_name).stem)),
        "domain": first_match(compact, [r"Домен\s*\n+([^\n]+)", r"Domain\s*[:\n]+([^\n]+)"], "unknown"),
        "load_frequency": first_match(compact, [r"Регламент загрузки\s*\n+([^\n]+)", r"Load frequency\s*[:\n]+([^\n]+)"], "unknown"),
        "load_type": first_match(compact, [r"Тип загрузки\s*\n+([^\n]+)", r"Load type\s*[:\n]+([^\n]+)"], "unknown"),
        "fields": fields,
        "source_mappings": [],
        "algorithms": [],
        "available_datetime_fields": datetime_fields,
        "supported_grains": ["day"] if datetime_fields else [],
        "supported_filters": [field["name"] for field in fields if field["name"].lower().endswith(("_nm", "_cd", "_id"))][:12],
        "dq_checks": [],
    }


def is_missing_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def as_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric if math.isfinite(numeric) else None
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if not text:
            return None
        try:
            numeric = float(text)
        except ValueError:
            return None
        return numeric if math.isfinite(numeric) else None
    return None


def looks_like_datetime_field(name: str) -> bool:
    lowered = name.lower()
    return lowered.endswith(("dt", "dttm", "date")) or "date" in lowered or "time" in lowered


def median(values: list[float]) -> float:
    ordered = sorted(values)
    count = len(ordered)
    mid = count // 2
    if count % 2:
        return ordered[mid]
    return (ordered[mid - 1] + ordered[mid]) / 2


def outlier_summary(values: list[float]) -> dict[str, Any]:
    if len(values) < 3:
        return {"count": 0, "low": None, "high": None}
    ordered = sorted(values)
    midpoint = len(ordered) // 2
    lower_half = ordered[:midpoint] or ordered
    upper_half = ordered[midpoint + (len(ordered) % 2) :] or ordered
    q1 = median(lower_half)
    q3 = median(upper_half)
    iqr = q3 - q1
    if iqr > 0:
        low = q1 - 1.5 * iqr
        high = q3 + 1.5 * iqr
    else:
        center = median(ordered)
        low = center
        high = center
    outliers = [value for value in ordered if value < low or value > high]
    if not outliers:
        center = median(ordered)
        if center != 0 and max(ordered) > abs(center) * 10:
            outliers = [max(ordered)]
            high = max(ordered)
        elif center != 0 and min(ordered) < -abs(center) * 10:
            outliers = [min(ordered)]
            low = min(ordered)
    return {"count": len(outliers), "low": low, "high": high, "examples": outliers[:5]}


def profile_rows(source_id: str, source_file: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    field_names: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for field in row:
            if field not in seen:
                seen.add(field)
                field_names.append(field)
    field_profiles: dict[str, dict[str, Any]] = {}
    for field in field_names:
        values = [row.get(field) for row in rows]
        null_count = sum(1 for value in values if is_missing_value(value))
        non_null_values = [value for value in values if not is_missing_value(value)]
        numeric_values: list[float] = []
        numeric_parse_failures = 0
        for value in non_null_values:
            parsed = as_float(value)
            if parsed is None:
                numeric_parse_failures += 1
            else:
                numeric_values.append(parsed)
        datetime_values: list[datetime] = []
        datetime_parse_failures = 0
        if looks_like_datetime_field(field):
            for value in non_null_values:
                parsed_dt = parse_datetime_value(value)
                if parsed_dt is None:
                    datetime_parse_failures += 1
                else:
                    datetime_values.append(parsed_dt)
        distinct_sample = []
        distinct_seen: set[str] = set()
        for value in non_null_values:
            text = str(value)
            if text not in distinct_seen:
                distinct_seen.add(text)
                distinct_sample.append(text)
            if len(distinct_sample) >= 25:
                break
        field_profiles[field] = {
            "field_name": field,
            "row_count": len(rows),
            "null_count": null_count,
            "null_ratio": round(null_count / len(rows), 4) if rows else 0,
            "numeric_parse_failures": numeric_parse_failures,
            "date_parse_failures": datetime_parse_failures,
            "numeric_min": min(numeric_values) if numeric_values else None,
            "numeric_max": max(numeric_values) if numeric_values else None,
            "datetime_min": compact_datetime(min(datetime_values)) if datetime_values else None,
            "datetime_max": compact_datetime(max(datetime_values)) if datetime_values else None,
            "distinct_count_sample": len(distinct_seen),
            "distinct_values_sample": distinct_sample,
            "outlier_summary": outlier_summary(numeric_values) if numeric_values else {"count": 0, "low": None, "high": None},
        }
    return {
        "profile_id": source_id,
        "source_file": source_file,
        "row_count": len(rows),
        "profiled_row_limit": MAX_SAMPLE_PROFILE_ROWS,
        "fields": list(field_profiles.values()),
        "field_names": field_names,
        "caveats": [] if len(rows) < MAX_SAMPLE_PROFILE_ROWS else ["profile limited to first 10000 rows"],
    }


def profile_csv_file(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for index, row in enumerate(reader):
            if index >= MAX_SAMPLE_PROFILE_ROWS:
                break
            rows.append(dict(row))
    return [profile_rows(slugify(path.stem), path.name, rows)]


def profile_xlsx_file(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to profile .xlsx sample files and is not available.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    profiles: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            profiles.append(profile_rows(f"{slugify(path.stem)}__{slugify(sheet.title)}", f"{path.name}:{sheet.title}", []))
            continue
        field_names = [str(header).strip() if header is not None else f"column_{index + 1}" for index, header in enumerate(headers)]
        rows: list[dict[str, Any]] = []
        for index, values in enumerate(rows_iter):
            if index >= MAX_SAMPLE_PROFILE_ROWS:
                break
            rows.append({field: values[column_index] if column_index < len(values) else None for column_index, field in enumerate(field_names)})
        profiles.append(profile_rows(f"{slugify(path.stem)}__{slugify(sheet.title)}", f"{path.name}:{sheet.title}", rows))
    return profiles


def profile_sample_file(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return profile_csv_file(path)
    if suffix == ".xlsx":
        return profile_xlsx_file(path)
    raise RuntimeError(f"Unsupported sample file type {suffix!r}; supported types are .csv and .xlsx.")


def dedupe_fields(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    output: list[dict[str, Any]] = []
    for field in fields:
        name = str(field.get("name") or "").strip()
        key = name.lower()
        if not name or key in seen:
            continue
        seen.add(key)
        output.append(field)
    return output


def default_metric_semantics() -> dict[str, Any]:
    return {
        "unit": "declared_or_pending",
        "grain": "declared_or_pending",
        "aggregation": "declared_or_pending",
        "numerator": "not_applicable_or_pending",
        "denominator": "not_applicable_or_pending",
        "additivity": "declared_or_pending",
        "time_grain": "declared_or_pending",
        "comparator": "declared_or_pending",
        "baseline": "declared_or_pending",
        "target": "declared_or_pending",
    }


def normalize_fixture_case(case: dict[str, Any]) -> dict[str, Any]:
    requirements = case.get("id_placeholder") or {}
    contracts = case.get("data_contracts") or []
    return {
        "case_id": str(case.get("case_id") or slugify(requirements.get("dashboard_name", "dashboard"))),
        "domain": case.get("domain", "unknown"),
        "source_manifest": case.get("source_manifest") or [],
        "id_placeholder": requirements,
        "data_contracts": contracts,
        "sample_profiles": case.get("sample_profiles") or [],
        "as_of": case.get("as_of"),
        "expected_outputs": case.get("expected_outputs") or {},
    }


def normalize_passport(case: dict[str, Any]) -> dict[str, Any]:
    requirements = case["id_placeholder"]
    return {
        "schema_name": "DashboardRequirementsPassport",
        "dashboard_name": requirements.get("dashboard_name", case["case_id"]),
        "contact": requirements.get("contact", "missing"),
        "process": requirements.get("process", "missing"),
        "lifetime": requirements.get("lifetime", "missing"),
        "update_frequency": requirements.get("update_frequency", "missing"),
        "objective": requirements.get("objective", "missing"),
        "background": requirements.get("background", "missing"),
        "business_value": requirements.get("business_value", "missing"),
        "audience": requirements.get("audience") or [],
        "decision_action": requirements.get("decision_action", "missing"),
        "data_sources": requirements.get("data_sources") or [],
        "id_placeholder": requirements.get("id_placeholder") or [],
        "source_statuses": requirements.get("source_statuses") or [],
        "metrics": requirements.get("metrics") or [],
        "calculation_methods": requirements.get("calculation_methods") or [],
        "visual_requirements": requirements.get("visual_requirements") or [],
        "sample_sketches": requirements.get("sample_sketches") or [],
        "open_questions": requirements.get("open_questions") or [],
        "out_of_scope": requirements.get("out_of_scope") or [],
        "missing_context": requirements.get("missing_context") or [],
        "manual_chart_type_required": False,
    }


def normalize_data_contracts(case: dict[str, Any]) -> list[dict[str, Any]]:
    contracts: list[dict[str, Any]] = []
    for raw in case.get("data_contracts") or []:
        if not isinstance(raw, dict):
            continue
        fields = dedupe_fields(raw.get("fields") or [])
        datetime_fields = raw.get("available_datetime_fields") or [
            field["name"] for field in fields if str(field.get("name", "")).lower().endswith(("dt", "dttm", "date"))
        ]
        contracts.append(
            {
                "schema_name": "DataContract",
                "contract_id": raw.get("contract_id") or slugify(raw.get("table_name", "contract")),
                "table_name": raw.get("table_name", "missing"),
                "domain": raw.get("domain", "missing"),
                "load_frequency": raw.get("load_frequency", "missing"),
                "load_type": raw.get("load_type", "missing"),
                "fields": fields,
                "source_mappings": raw.get("source_mappings") or [],
                "algorithms": raw.get("algorithms") or [],
                "available_datetime_fields": datetime_fields,
                "supported_grains": raw.get("supported_grains") or [],
                "supported_filters": raw.get("supported_filters") or [],
                "dq_checks": raw.get("dq_checks") or [],
            }
        )
    return contracts


def field_index(data_contracts: list[dict[str, Any]]) -> set[str]:
    fields: set[str] = set()
    for contract in data_contracts:
        for field in contract.get("fields") or []:
            fields.add(str(field.get("name", "")).lower())
    return fields


def field_contract_index(data_contracts: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    for contract in data_contracts:
        for field in contract.get("fields") or []:
            if field.get("name"):
                fields[str(field["name"]).lower()] = field
    return fields


def profile_field_index(sample_profiles: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    for profile in sample_profiles:
        for field in profile.get("fields") or []:
            name = str(field.get("field_name") or "").lower()
            if name and name not in fields:
                fields[name] = {**field, "profile_id": profile.get("profile_id"), "source_file": profile.get("source_file")}
    return fields


def make_finding(code: str, severity: str, message: str, *, fields: list[str] | None = None, gate: str | None = None) -> dict[str, Any]:
    finding: dict[str, Any] = {
        "code": code,
        "severity": severity,
        "message": message,
    }
    if fields:
        finding["fields"] = fields
    if gate:
        finding["gate"] = gate
    return finding


def metric_is_percent_share_or_rate(semantics: dict[str, Any]) -> bool:
    text = " ".join(str(semantics.get(key, "")) for key in ("unit", "aggregation", "numerator", "denominator", "additivity")).lower()
    return any(term in text for term in ("percent", "percentage", "share", "rate", "ratio", "conversion"))


def denominator_is_pending(semantics: dict[str, Any]) -> bool:
    denominator = str(semantics.get("denominator", "")).strip().lower()
    return not denominator or denominator in {"pending", "unknown", "missing", "not_applicable_or_pending", "declared_or_pending", "not_applicable"}


def denominator_field_missing(semantics: dict[str, Any], available_fields: set[str]) -> bool:
    denominator = str(semantics.get("denominator", "")).strip()
    if not denominator or denominator.lower() in {"pending", "unknown", "missing", "not_applicable_or_pending", "declared_or_pending", "not_applicable"}:
        return False
    tokens = [token for token in re.split(r"[^A-Za-z0-9_]+", denominator) if "_" in token]
    return any(token.lower() not in available_fields for token in tokens)


def additivity_pending(semantics: dict[str, Any]) -> bool:
    additivity = str(semantics.get("additivity", "")).lower()
    return not additivity or additivity in {"pending", "unknown", "declared_or_pending"} or "pending" in additivity


def additivity_non_additive(semantics: dict[str, Any]) -> bool:
    additivity = str(semantics.get("additivity", "")).lower()
    return "non_additive" in additivity or additivity == "mixed"


def selected_family_from_metric(metric: dict[str, Any]) -> str:
    return str(metric.get("expected_family") or metric.get("family") or "")


def metric_time_fields(metric: dict[str, Any], contract_fields: dict[str, dict[str, Any]]) -> list[str]:
    fields = []
    for field in metric.get("required_fields") or []:
        field_name = str(field)
        contract = contract_fields.get(field_name.lower()) or {}
        flags = {str(flag).lower() for flag in contract.get("flags") or []}
        if "date_time" in flags or looks_like_datetime_field(field_name):
            fields.append(field_name)
    return fields


def freshness_threshold_hours(update_frequency: str) -> int:
    text = update_frequency.lower()
    for key, hours in FRESHNESS_DEFAULT_HOURS.items():
        if key != "unknown" and key in text:
            return hours
    return FRESHNESS_DEFAULT_HOURS["unknown"]


def customer_state_for_status(status: str) -> str:
    if status in {"data_missing", "source_not_integrated"}:
        return "SOURCE MISSING"
    if status == "quality_risk":
        return "QUALITY ISSUE"
    if status == "methodology_pending":
        return "METHODOLOGY PENDING"
    return "AVAILABLE"


def recommended_fix(status: str, quality_findings: list[dict[str, Any]], methodology_findings: list[dict[str, Any]], metric: dict[str, Any]) -> str:
    if metric.get("recommended_fix_or_question"):
        return str(metric["recommended_fix_or_question"])
    if metric.get("methodology_pending_action"):
        return str(metric["methodology_pending_action"])
    findings = quality_findings + methodology_findings
    if findings:
        first = findings[0]
        if status in {"data_missing", "source_not_integrated"}:
            return "Provide or integrate the missing source fields before finalizing this widget."
        if status == "quality_risk":
            return f"Resolve data quality issue: {first['message']}"
        if status == "methodology_pending":
            return f"Confirm methodology decision: {first['message']}"
    return "none"


def evaluate_quality_and_methodology(
    metric: dict[str, Any],
    *,
    data_contracts: list[dict[str, Any]],
    sample_profiles: list[dict[str, Any]],
    as_of: datetime | None,
    update_frequency: str,
    matched_fields: list[str],
    missing_fields: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str | None]:
    quality_findings: list[dict[str, Any]] = []
    methodology_findings: list[dict[str, Any]] = []
    available_fields = field_index(data_contracts)
    contract_fields = field_contract_index(data_contracts)
    profile_fields = profile_field_index(sample_profiles)
    semantics = {**default_metric_semantics(), **(metric.get("metric_semantics") or {})}
    family = selected_family_from_metric(metric)
    task = str(metric.get("analytical_task") or "").lower()

    if missing_fields:
        quality_findings.append(
            make_finding(
                "missing_s2t_field",
                "fail",
                "Required S2T fields are missing: " + ", ".join(missing_fields),
                fields=missing_fields,
                gate="DQ-F003",
            )
        )
    if metric.get("source_status") == "not_integrated" or metric.get("support_status") == "source_not_integrated":
        quality_findings.append(make_finding("source_not_integrated", "fail", "Required source is named but not integrated.", gate="DQ-F001"))

    for field in matched_fields:
        profile = profile_fields.get(field.lower())
        if not profile:
            continue
        contract_field = contract_fields.get(field.lower()) or {}
        contract_type = str(contract_field.get("type", "")).lower()
        contract_flags = {str(flag).lower() for flag in contract_field.get("flags") or []}
        field_lower = field.lower()
        numeric_expected = (
            contract_type in {"number", "numeric", "float", "double", "integer", "int", "decimal"}
            or "metric" in contract_flags
            or field_lower.endswith(("_cnt", "_qty", "_pct", "_amount", "_num", "_rate"))
        )
        null_ratio = float(profile.get("null_ratio") or 0)
        null_count = int(profile.get("null_count") or 0)
        if null_count > 0:
            severity = "fail" if null_ratio >= NULL_RATIO_QUALITY_FAIL else "warn"
            quality_findings.append(
                make_finding(
                    "null_values",
                    severity,
                    f"{field} has {null_count} null values ({null_ratio:.0%}).",
                    fields=[field],
                    gate="DQ-F012" if severity == "fail" else "DQ-W004",
                )
            )
        if numeric_expected and int(profile.get("numeric_parse_failures") or 0) > 0 and not looks_like_datetime_field(field):
            quality_findings.append(
                make_finding(
                    "invalid_numeric_values",
                    "fail",
                    f"{field} has values that cannot be parsed as numeric.",
                    fields=[field],
                    gate="DQ-F012",
                )
            )
        if int(profile.get("date_parse_failures") or 0) > 0:
            quality_findings.append(
                make_finding(
                    "invalid_datetime_values",
                    "fail",
                    f"{field} has values that cannot be parsed as datetime.",
                    fields=[field],
                    gate="DQ-F012",
                )
            )
        outliers = profile.get("outlier_summary") or {}
        if int(outliers.get("count") or 0) > 0:
            quality_findings.append(
                make_finding(
                    "outliers_detected",
                    "warn",
                    f"{field} has potential outliers by IQR/range check.",
                    fields=[field],
                    gate="DQ-W001",
                )
            )

    time_fields = metric_time_fields(metric, contract_fields)
    if as_of and time_fields:
        threshold_hours = freshness_threshold_hours(update_frequency)
        for field in time_fields:
            profile = profile_fields.get(field.lower())
            if not profile or not profile.get("datetime_max"):
                continue
            max_dt = parse_datetime_value(profile.get("datetime_max"))
            if max_dt and (as_of - max_dt).total_seconds() > threshold_hours * 3600:
                quality_findings.append(
                    make_finding(
                        "stale_data",
                        "fail",
                        f"{field} max timestamp {compact_datetime(max_dt)} is older than freshness threshold {threshold_hours}h.",
                        fields=[field],
                        gate="DQ-F002",
                    )
                )

    if metric_is_percent_share_or_rate(semantics):
        if denominator_field_missing(semantics, available_fields):
            denominator = str(semantics.get("denominator"))
            quality_findings.append(
                make_finding(
                    "denominator_field_missing",
                    "fail",
                    f"Denominator field is absent from S2T: {denominator}.",
                    gate="DQ-F004",
                )
            )
        elif denominator_is_pending(semantics):
            methodology_findings.append(make_finding("denominator_pending", "fail", "Percent/rate/share metric needs denominator proof.", gate="DQ-F004"))

    if family in {"stacked_100", "waterfall"} or task == "part_to_whole":
        if additivity_non_additive(semantics):
            quality_findings.append(make_finding("non_additive_share_or_stack", "fail", f"{family or task} requires additive components.", gate="DQ-F008"))
        elif additivity_pending(semantics):
            methodology_findings.append(make_finding("additivity_pending", "fail", f"{family or task} requires additivity proof.", gate="DQ-F008"))

    if task in {"id_placeholder", "id_placeholder"} or family in {"histogram", "box_plot", "scatter", "bubble"}:
        row_count = max([int(profile.get("row_count") or 0) for profile in sample_profiles] or [0])
        if row_count:
            if row_count < SMALL_SAMPLE_FAIL:
                quality_findings.append(
                    make_finding(
                        "sample_size_too_small",
                        "fail",
                        f"Sample size {row_count} is below minimum {SMALL_SAMPLE_FAIL} for final distribution/relationship claim.",
                        gate="DIST-F002" if task == "id_placeholder" else "UNC-F001",
                    )
                )
            elif row_count < SMALL_SAMPLE_WARN:
                quality_findings.append(
                    make_finding(
                        "sample_size_warning",
                        "warn",
                        f"Sample size {row_count} is below preferred {SMALL_SAMPLE_WARN}; add caveat.",
                        gate="DQ-W004",
                    )
                )
        else:
            methodology_findings.append(make_finding("sample_size_unverified", "warn", "Sample size was not verified from raw data profile.", gate="DQ-F005"))
    if task == "id_placeholder" or family in {"scatter", "bubble"}:
        if not metric.get("causal_design_evidence"):
            methodology_findings.append(
                make_finding(
                    "correlation_causation_caveat",
                    "warn",
                    "Relationship view must not imply causation without causal design evidence.",
                    gate="DQ-W007",
                )
            )

    if task in {"time_trend", "time_bucket", "seasonality"} or family in {"line_chart", "vertical_bar_time_bucket", "combo_time_series_combo", "bump_chart"}:
        time_grain = str(semantics.get("time_grain", "")).strip().lower()
        if not time_grain or "pending" in time_grain or time_grain in {"declared_or_pending", "unknown"}:
            methodology_findings.append(make_finding("time_grain_pending", "fail", "Trend claim requires a declared time grain.", gate="TIME-F001"))
        time_bucket_count = 0
        for field in time_fields:
            profile = profile_fields.get(field.lower())
            if profile:
                time_bucket_count = max(time_bucket_count, int(profile.get("distinct_count_sample") or 0))
        if 0 < time_bucket_count < 4:
            methodology_findings.append(
                make_finding(
                    "too_few_time_points",
                    "fail",
                    f"Trend claim has only {time_bucket_count} observed time points; at least 4 are required.",
                    gate="TIME-F004",
                )
            )

    explicit_quality_risks = [str(risk) for risk in metric.get("quality_risks") or []]
    for risk in explicit_quality_risks:
        quality_findings.append(make_finding("declared_quality_risk", "fail", risk, gate="DQ-F012"))
    for item in metric.get("missing_context") or []:
        if metric.get("support_status") == "methodology_pending":
            methodology_findings.append(make_finding("declared_methodology_gap", "fail", str(item)))

    status_override: str | None = None
    if any(finding["code"] == "source_not_integrated" for finding in quality_findings):
        status_override = "source_not_integrated"
    elif missing_fields or any(finding["code"] == "denominator_field_missing" for finding in quality_findings):
        status_override = "data_missing"
    elif any(finding["severity"] == "fail" for finding in quality_findings):
        status_override = "quality_risk"
    elif any(finding["severity"] == "fail" for finding in methodology_findings):
        status_override = "methodology_pending"
    return quality_findings, methodology_findings, status_override


def classify_metric(
    metric: dict[str, Any],
    data_contracts: list[dict[str, Any]],
    *,
    sample_profiles: list[dict[str, Any]] | None = None,
    as_of: datetime | None = None,
    update_frequency: str = "unknown",
) -> dict[str, Any]:
    required_fields = [str(field) for field in metric.get("required_fields") or []]
    available_fields = field_index(data_contracts)
    matched = [field for field in required_fields if field.lower() in available_fields]
    missing = [field for field in required_fields if field.lower() not in available_fields]
    status = metric.get("support_status")
    if status not in SUPPORT_STATUSES:
        if required_fields and not missing:
            status = "supported_with_assumption" if metric.get("assumptions") else "supported"
        elif metric.get("source_status") == "not_integrated":
            status = "source_not_integrated"
        elif metric.get("methodology_pending") or metric.get("missing_context"):
            status = "methodology_pending"
        elif missing:
            status = "data_missing"
        else:
            status = "supported_with_assumption"
    quality_findings, methodology_findings, status_override = evaluate_quality_and_methodology(
        metric,
        data_contracts=data_contracts,
        sample_profiles=sample_profiles or [],
        as_of=as_of,
        update_frequency=update_frequency,
        matched_fields=matched,
        missing_fields=missing,
    )
    if status_override:
        status = status_override
    availability = metric.get("availability_state") or STATUS_TO_AVAILABILITY[status]
    customer_visible_state = customer_state_for_status(status)
    return {
        "metric_id": metric["metric_id"],
        "metric_name": metric.get("name", metric["metric_id"]),
        "business_question": metric.get("business_question", ""),
        "analytical_task": metric.get("analytical_task", "kpi_status"),
        "support_status": status,
        "data_support_status": status,
        "availability_state": availability,
        "customer_visible_state": customer_visible_state,
        "required_fields": required_fields,
        "matched_fields": matched,
        "missing_fields": missing,
        "source_contract_ids": metric.get("source_contract_ids") or [],
        "assumptions": metric.get("assumptions") or [],
        "missing_context": metric.get("missing_context") or [],
        "quality_risks": metric.get("quality_risks") or [],
        "quality_findings": quality_findings,
        "methodology_findings": methodology_findings,
        "recommended_fix_or_question": recommended_fix(status, quality_findings, methodology_findings, metric),
        "metric_semantics": {**default_metric_semantics(), **(metric.get("metric_semantics") or {})},
    }


def infer_family(metric: dict[str, Any], support: dict[str, Any]) -> str:
    requested = metric.get("expected_family") or metric.get("family")
    if requested:
        return str(requested)
    task = str(metric.get("analytical_task") or "").lower()
    name = f"{metric.get('name', '')} {metric.get('business_question', '')}".lower()
    if task == "kpi_status" or "uptime" in name:
        return "kpi_value_sparkline"
    if task in {"time_trend", "seasonality"}:
        return "line_chart"
    if task in {"time_bucket", "delivery_rate"}:
        return "vertical_bar_time_bucket"
    if task in {"flow_or_process", "conversion"}:
        return "funnel_snapshot"
    if task in {"ranking", "category_comparison"}:
        return "horizontal_bar"
    if task == "matrix":
        return "heatmap"
    if task == "id_placeholder":
        return "box_plot"
    if task == "id_placeholder":
        return "bubble" if "size" in support["metric_semantics"] else "scatter"
    if task == "geo_pattern":
        return "native_map_geo_widget"
    if task == "exact_lookup":
        return "table_node"
    if task in {"metadata", "methodology_or_metadata"}:
        return "md_methodology_block"
    return "table_node" if support["support_status"] in {"data_missing", "source_not_integrated"} else "kpi_value_sparkline"


def infer_chart_decision(metric: dict[str, Any], support: dict[str, Any]) -> dict[str, Any]:
    family = infer_family(metric, support)
    if family not in ALLOWED_VISUAL_FAMILIES:
        family = "table_node"
    route = metric.get("expected_route") or expected_route_for_family(family)
    tabs = expected_tabs_for_route(route)
    status = STATUS_TO_DECISION_STATUS[support["support_status"]]
    return {
        "decision_id": f"{support['metric_id']}__chart_decision",
        "metric_id": support["metric_id"],
        "selected_family": family,
        "family_contract_id": family,
        "selected_variant": metric.get("expected_variant") or family,
        "route": route,
        "required_editor_tabs": tabs,
        "analytical_task": support["analytical_task"],
        "business_question": support["business_question"],
        "requirement_phrase_evidence": metric.get("requirement_phrase", metric.get("name", support["metric_id"])),
        "s2t_data_shape_evidence": {
            "required_fields": support["required_fields"],
            "matched_fields": support["matched_fields"],
            "missing_fields": support["missing_fields"],
        },
        "metric_semantics_evidence": support["metric_semantics"],
        "simpler_alternative_considered": metric.get("simpler_alternative", "table_node"),
        "split_widget_decision": metric.get("split_widget_rule", "single_widget_unless_metric_grain_or_audience_differs"),
        "rejected_alternatives": metric.get("rejected_alternatives") or ["manual_chart_type_prompt", "decorative_visual"],
        "assumptions": support["assumptions"],
        "missing_context": support["missing_context"],
        "availability_state": support["availability_state"],
        "data_support_status": support["data_support_status"],
        "quality_findings": support["quality_findings"],
        "methodology_findings": support["methodology_findings"],
        "customer_visible_state": support["customer_visible_state"],
        "recommended_fix_or_question": support["recommended_fix_or_question"],
        "status": status,
        "confidence": metric.get("confidence") or ("high" if support["support_status"] == "supported" else "medium"),
        "geo_evidence": metric.get("geo_evidence"),
        "safe_api_notes": [
            "governance_output_only",
            "datalens-api-integration_consumes_preapproved_spec",
            "no_live_write_from_intake",
        ],
    }


def customer_message(metric: dict[str, Any], support: dict[str, Any]) -> str:
    if metric.get("customer_visible_message"):
        return str(metric["customer_visible_message"])
    state = support["availability_state"]
    if state == "available":
        return "Metric is available for baseline dashboard generation."
    if state == "unavailable_methodology_pending":
        return "Planned widget is visible, but the calculation method needs confirmation."
    if state == "unavailable_invalid_data":
        return "Planned widget is visible, but data quality must be resolved before using the value."
    return "Planned widget is visible, but required source data is unavailable or not integrated."


def build_visual_spec(metric: dict[str, Any], support: dict[str, Any], decision: dict[str, Any]) -> dict[str, Any]:
    return {
        "visual_id": f"{support['metric_id']}__visual",
        "metric_id": support["metric_id"],
        "title": metric.get("name", support["metric_id"]),
        "family": decision["selected_family"],
        "route": decision["route"],
        "required_editor_tabs": decision["required_editor_tabs"],
        "availability_state": support["availability_state"],
        "state_title": STATE_TITLES.get(support["availability_state"], "NO DATA"),
        "data_support_status": support["data_support_status"],
        "quality_findings": support["quality_findings"],
        "methodology_findings": support["methodology_findings"],
        "customer_visible_state": support["customer_visible_state"],
        "recommended_fix_or_question": support["recommended_fix_or_question"],
        "missing_or_invalid_reason": "; ".join(support["missing_fields"] or support["missing_context"] or support["quality_risks"] or ["none"]),
        "customer_visible_message": customer_message(metric, support),
        "methodology_pending_action": metric.get("methodology_pending_action")
        or ("Confirm calculation method and numerator/denominator." if support["support_status"] == "methodology_pending" else "none"),
        "style_contract": {
            "labels": metric.get("label_policy", "direct_labels_where_readable"),
            "legend": metric.get("legend_policy", "remove_when_direct_labels_suffice"),
            "gridlines": metric.get("gridline_policy", "subtle_only_when_scale_reading_matters"),
            "color": metric.get("color_policy", "neutral_context_with_semantic_focus_or_alert_only"),
            "tooltip": metric.get("tooltip_policy", "show metric definition, unit, grain, freshness, and source"),
            "empty_state": "use required availability_state title and short reason",
        },
    }


def build_kpi_registry(passport: dict[str, Any], metric_support: list[dict[str, Any]]) -> dict[str, Any]:
    metrics = []
    passport_metrics = {metric["metric_id"]: metric for metric in passport.get("metrics") or [] if isinstance(metric, dict) and "metric_id" in metric}
    for support in metric_support:
        source = passport_metrics.get(support["metric_id"], {})
        metrics.append(
            {
                "metric_id": support["metric_id"],
                "name": support["metric_name"],
                "business_question": support["business_question"],
                "semantics": support["metric_semantics"],
                "support_status": support["support_status"],
                "availability_state": support["availability_state"],
                "calculation_method": source.get("calculation_method", "declared_or_pending"),
            }
        )
    return {
        "schema_name": "KPIRegistry",
        "dashboard_name": passport["dashboard_name"],
        "metrics": metrics,
    }


def build_dashboard_build_spec(passport: dict[str, Any], decisions: list[dict[str, Any]], visuals: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "schema_name": "DashboardBuildSpec",
        "dashboard_name": passport["dashboard_name"],
        "scope_policy": "preserve_full_scope_with_unavailable_states",
        "reading_path": "primary signal -> diagnosis -> detail/action -> metadata/freshness",
        "selectors": [
            {
                "selector_id": "global_period_selector",
                "family": "date_range_selector",
                "route": "editor_js_control",
                "required_editor_tabs": expected_tabs_for_route("editor_js_control"),
            }
        ],
        "sections": [
            {
                "section_id": "overview",
                "title": "Overview",
                "visual_ids": [visual["visual_id"] for visual in visuals[:3]],
            },
            {
                "section_id": "diagnostics",
                "title": "Diagnostics",
                "visual_ids": [visual["visual_id"] for visual in visuals[3:]],
            },
            {
                "section_id": "metadata",
                "title": "Metadata and methodology",
                "visual_ids": ["metadata_methodology__visual"],
            },
        ],
        "chart_decision_ids": [decision["decision_id"] for decision in decisions],
        "visual_ids": [visual["visual_id"] for visual in visuals],
        "availability_summary": {
            "available": sum(1 for visual in visuals if visual["availability_state"] == "available"),
            "unavailable": sum(1 for visual in visuals if visual["availability_state"] != "available"),
        },
    }


def render_dashboard_plan(bundle: dict[str, Any]) -> str:
    passport = bundle["dashboard_requirements_passport"]
    lines = [
        f"# {passport['dashboard_name']} Dashboard Plan",
        "",
        f"- Case ID: `{bundle['case_id']}`",
        f"- Objective: {passport.get('objective', 'missing')}",
        f"- Audience: {', '.join(passport.get('audience') or ['missing'])}",
        f"- Decision/action: {passport.get('decision_action', 'missing')}",
        f"- Scope policy: `{bundle['dashboard_build_spec']['scope_policy']}`",
        "",
        "## Metric Support",
        "",
        "| Metric | data_support_status | customer_visible_state | Family | Route | quality_findings | methodology_findings | recommended_fix_or_question |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    decisions = {decision["metric_id"]: decision for decision in bundle["chart_decisions"]}
    for support in bundle["metric_support_matrix"]:
        decision = decisions[support["metric_id"]]
        quality_codes = ", ".join(finding["code"] for finding in support.get("quality_findings") or []) or "none"
        methodology_codes = ", ".join(finding["code"] for finding in support.get("methodology_findings") or []) or "none"
        lines.append(
            f"| {support['metric_name']} | `{support['data_support_status']}` | `{support['customer_visible_state']}` | `{decision['selected_family']}` | `{decision['route']}` | `{quality_codes}` | `{methodology_codes}` | {support['recommended_fix_or_question']} |"
        )
    lines.extend(["", "## Open Questions", ""])
    questions = passport.get("open_questions") or ["none"]
    lines.extend(f"- {question}" for question in questions)
    return "\n".join(lines) + "\n"


def render_memory_context(bundle: dict[str, Any]) -> str:
    passport = bundle["dashboard_requirements_passport"]
    status_counts: dict[str, int] = {}
    for support in bundle["metric_support_matrix"]:
        status_counts[support["support_status"]] = status_counts.get(support["support_status"], 0) + 1
    source_names = [source.get("file_name", source.get("source_id", "unknown")) for source in bundle.get("source_manifest") or []]
    lines = [
        f"# Requirements/S2T Context: {passport['dashboard_name']}",
        "",
        f"- Case ID: `{bundle['case_id']}`",
        f"- Source roles: {', '.join(source_names) if source_names else 'fixture evidence only'}",
        f"- Update frequency: {passport.get('update_frequency', 'missing')}",
        f"- Decision/action: {passport.get('decision_action', 'missing')}",
        f"- Manual chart type required: `{passport.get('manual_chart_type_required')}`",
        f"- Metric status counts: `{json.dumps(status_counts, ensure_ascii=False, sort_keys=True)}`",
        "",
        "## Next Context For Codex",
        "",
        "- Use this bundle as the dashboard thread memory before implementation.",
        "- Generate all planned widgets; unavailable metrics keep explicit customer-facing states.",
        "- DataLens API integration remains downstream-only and consumes approved chart decisions.",
    ]
    return "\n".join(lines) + "\n"


def build_governance_bundle(case: dict[str, Any], *, sample_profiles: list[dict[str, Any]] | None = None, as_of: datetime | str | None = None) -> dict[str, Any]:
    normalized = normalize_fixture_case(case)
    passport = normalize_passport(normalized)
    data_contracts = normalize_data_contracts(normalized)
    all_sample_profiles = [*(normalized.get("sample_profiles") or []), *(sample_profiles or [])]
    if isinstance(as_of, str):
        as_of_dt = parse_as_of(as_of)
    elif isinstance(as_of, datetime):
        as_of_dt = as_of
    else:
        as_of_dt = parse_as_of(normalized.get("as_of"))
    metric_support: list[dict[str, Any]] = []
    decisions: list[dict[str, Any]] = []
    visuals: list[dict[str, Any]] = []
    for metric in passport.get("metrics") or []:
        if not isinstance(metric, dict) or not metric.get("metric_id"):
            continue
        support = classify_metric(metric, data_contracts, sample_profiles=all_sample_profiles, as_of=as_of_dt, update_frequency=passport.get("update_frequency", "unknown"))
        decision = infer_chart_decision(metric, support)
        visual = build_visual_spec(metric, support, decision)
        metric_support.append(support)
        decisions.append(decision)
        visuals.append(visual)
    metadata_visual = {
        "visual_id": "metadata_methodology__visual",
        "metric_id": "metadata_methodology",
        "title": "Metadata and methodology",
        "family": "md_methodology_block",
        "route": "editor_markdown",
        "required_editor_tabs": expected_tabs_for_route("editor_markdown"),
        "availability_state": "available",
        "state_title": "NO DATA",
        "data_support_status": "supported",
        "quality_findings": [],
        "methodology_findings": [],
        "customer_visible_state": "AVAILABLE",
        "recommended_fix_or_question": "none",
        "missing_or_invalid_reason": "none",
        "customer_visible_message": "Shows source, freshness, owner, and pending methodology notes.",
        "methodology_pending_action": "Keep pending metric questions visible until resolved.",
        "style_contract": {
            "labels": "plain_text",
            "legend": "not_applicable",
            "gridlines": "not_applicable",
            "color": "neutral_context_only",
            "tooltip": "not_applicable",
            "empty_state": "not_applicable",
        },
    }
    visuals_with_metadata = [*visuals, metadata_visual]
    bundle = {
        "schema_version": SCHEMA_VERSION,
        "case_id": normalized["case_id"],
        "domain": normalized["domain"],
        "source_manifest": normalized.get("source_manifest") or [],
        "input_evidence_set": {
            "schema_name": "InputEvidenceSet",
            "requirements_sources": [source for source in normalized.get("source_manifest") or [] if "requirement" in str(source.get("role", ""))],
            "s2t_sources": [source for source in normalized.get("source_manifest") or [] if "s2t" in str(source.get("role", "")) or "contract" in str(source.get("role", ""))],
            "raw_data_profiles": [source for source in normalized.get("source_manifest") or [] if "raw" in str(source.get("role", ""))],
            "project_evidence": [],
        },
        "dashboard_requirements_passport": passport,
        "data_contracts": data_contracts,
        "sample_profiles": all_sample_profiles,
        "as_of": compact_datetime(as_of_dt),
        "metric_support_matrix": metric_support,
        "kpi_registry": build_kpi_registry(passport, metric_support),
        "chart_decisions": decisions,
        "dashboard_build_spec": build_dashboard_build_spec(passport, decisions, visuals_with_metadata),
        "visual_build_specs": visuals_with_metadata,
        "downstream_api_notes": {
            "role": "downstream_only",
            "requires_preapproved_governance_output": True,
            "live_writes": "not_performed_by_intake",
        },
    }
    bundle["dashboard_plan_markdown"] = render_dashboard_plan(bundle)
    bundle["memory_context_markdown"] = render_memory_context(bundle)
    return bundle


def case_from_pdfs(case_id: str, requirements_pdfs: list[Path], s2t_pdfs: list[Path], sample_files: list[Path] | None = None) -> dict[str, Any]:
    if not requirements_pdfs:
        raise RuntimeError("At least one --requirements-pdf is required in direct PDF mode.")
    requirement_texts = []
    source_manifest = []
    for path in requirements_pdfs:
        text = extract_pdf_text(path)
        requirement_texts.append(text)
        source_manifest.append({"source_id": slugify(path.stem), "role": "customer_requirement_pdf", "file_name": path.name, "sha256": sha256_file(path)})
    parsed_requirements = parse_requirement_text("\n\n".join(requirement_texts), requirements_pdfs[0].name)
    contracts = []
    for path in s2t_pdfs:
        text = extract_pdf_text(path)
        contracts.append(parse_s2t_text(text, path.name))
        source_manifest.append({"source_id": slugify(path.stem), "role": "s2t_data_contract_pdf", "file_name": path.name, "sha256": sha256_file(path)})
    sample_profiles: list[dict[str, Any]] = []
    for path in sample_files or []:
        sample_profiles.extend(profile_sample_file(path))
        source_manifest.append({"source_id": slugify(path.stem), "role": "real_data_sample", "file_name": path.name, "sha256": sha256_file(path)})
    return {
        "case_id": case_id,
        "domain": slugify(parsed_requirements.get("dashboard_name", case_id)),
        "source_manifest": source_manifest,
        "id_placeholder": parsed_requirements,
        "data_contracts": contracts,
        "sample_profiles": sample_profiles,
    }


def write_outputs(bundle: dict[str, Any], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "governance_bundle.json", {key: value for key, value in bundle.items() if not key.endswith("_markdown")})
    write_json(output_dir / "dashboard_requirements_passport.json", bundle["dashboard_requirements_passport"])
    write_json(output_dir / "data_contracts.json", bundle["data_contracts"])
    write_json(output_dir / "metric_support_matrix.json", bundle["metric_support_matrix"])
    (output_dir / "dashboard_plan.md").write_text(bundle["dashboard_plan_markdown"], encoding="utf-8")
    (output_dir / "memory_context.md").write_text(bundle["memory_context_markdown"], encoding="utf-8")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a compact governed dashboard intake bundle from requirements and S2T evidence.")
    parser.add_argument("--fixture", type=Path, help="Compact fixture case JSON.")
    parser.add_argument("--requirements-pdf", type=Path, action="append", default=[], help="Customer requirement PDF path; repeatable.")
    parser.add_argument("--s2t-pdf", type=Path, action="append", default=[], help="S2T/data-contract PDF path; repeatable.")
    parser.add_argument("--sample-file", type=Path, action="append", default=[], help="Optional local CSV/XLSX real data sample; repeatable.")
    parser.add_argument("--as-of", help="Deterministic ISO datetime for freshness checks.")
    parser.add_argument("--case-id", default="requirements_s2t_case", help="Case id for direct PDF mode.")
    parser.add_argument("--output-dir", type=Path, help="Directory for bundle artifacts.")
    parser.add_argument("--print-json", action="store_true", help="Print governance bundle JSON to stdout.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    try:
        if args.fixture:
            case = load_json(args.fixture)
            sample_profiles: list[dict[str, Any]] = []
            for sample_path in args.sample_file:
                sample_profiles.extend(profile_sample_file(sample_path))
                case.setdefault("source_manifest", []).append(
                    {"source_id": slugify(sample_path.stem), "role": "real_data_sample", "file_name": sample_path.name, "sha256": sha256_file(sample_path)}
                )
        else:
            case = case_from_pdfs(args.case_id, args.requirements_pdf, args.s2t_pdf, args.sample_file)
            sample_profiles = []
        bundle = build_governance_bundle(case, sample_profiles=sample_profiles, as_of=args.as_of)
        if args.output_dir:
            write_outputs(bundle, args.output_dir)
        if args.print_json:
            print(json.dumps(bundle, ensure_ascii=False, indent=2))
    except Exception as exc:  # noqa: BLE001
        print(f"requirements_s2t_intake: {exc}", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
